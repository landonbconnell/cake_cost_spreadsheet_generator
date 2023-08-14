from openpyxl import Workbook
from openpyxl.styles import NamedStyle

def displayIngredientData(ingredients):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ingredient Data"
    data = [["Ingredient", "Price", "Size (oz)", "Price Per Unit"]]

    for ingredient in ingredients:
        price_per_unit = float(ingredient["price"]) / float(ingredient["size"])
        data.append([ingredient["name"], ingredient["price"], ingredient["size"], price_per_unit])

    for row in data:
        ws.append(row)
    
    # Define the currency style
    currency_style = NamedStyle(name="currency_style", number_format="$0.00")
    
    # Add named style to the workbook if it doesn't exist
    if "currency_style" not in wb.named_styles:
        wb.add_named_style(currency_style)
    
    # Set the currency format for the "price" and "price per unit" columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):  # adjusting for "price" column
        for cell in row:
            cell.style = "currency_style"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):  # adjusting for "price per unit" column
        for cell in row:
            cell.style = "currency_style"
    
    wb.save("ingredientData.xlsx")
